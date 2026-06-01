from __future__ import annotations

import argparse
import json
import re
import warnings
from bisect import bisect_left, bisect_right
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation, getcontext
from pathlib import Path
from typing import Iterable

import openpyxl


getcontext().prec = 28
warnings.filterwarnings("ignore", message="Workbook contains no default style.*")

ROOT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_DIR = ROOT_DIR / "data" / "raw_label"
DEFAULT_OUTPUT = ROOT_DIR / "data" / "label" / "all" / "label_tags.json"

BINANCE_TIME_OFFSET = timedelta(hours=8)
MATCH_WINDOW = timedelta(days=120)
MAX_NEARBY_CANDIDATES = 30
MAX_SEQUENCE_LENGTH = 6
MAX_GROUPS_PER_ANCHOR = 3

STABLECOINS = {"USDT", "USDC", "BUSD", "FDUSD", "TUSD", "DAI", "PYUSD"}
KNOWN_FIAT = {
    "CNY",
    "USD",
    "EUR",
    "GBP",
    "HKD",
    "SGD",
    "JPY",
    "KRW",
    "RUB",
    "TRY",
    "AUD",
    "CAD",
    "THB",
    "VND",
    "INR",
}

MATCH_CONFIG = {
    "fiat_buy": {
        "counter_direction": "chain_withdrawal",
        "mode": "forward",
        "label_type": "fiat_buy_to_chain_withdrawal",
    },
    "chain_withdrawal": {
        "counter_direction": "fiat_buy",
        "mode": "backward",
        "label_type": "chain_withdrawal_to_fiat_buy",
    },
    "fiat_sell": {
        "counter_direction": "chain_deposit",
        "mode": "backward",
        "label_type": "fiat_sell_to_chain_deposit",
    },
    "chain_deposit": {
        "counter_direction": "fiat_sell",
        "mode": "forward",
        "label_type": "chain_deposit_to_fiat_sell",
    },
}


@dataclass
class Event:
    exchange: str
    source_file: str
    source_sheet: str
    source_row: int
    event_class: str
    direction: str
    asset: str
    quantity: Decimal
    time: datetime
    raw_time: str
    account_id: str | None = None
    fiat_currency: str | None = None
    fiat_amount: Decimal | None = None
    status: str | None = None
    txid: str | None = None
    address: str | None = None
    counterparty: str | None = None
    onchain: bool = True
    match_eligible: bool = True
    raw_direction: str | None = None
    note: str | None = None
    event_id: str | None = None


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    text = str(value)
    text = text.replace("\u3000", " ").replace("\xa0", " ").strip()
    if text.lower() == "null":
        return ""
    return text


def normalize_key(text: str) -> str:
    text = clean_text(text)
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    return text.lower()


def parse_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    text = clean_text(value).replace(",", "").strip("'")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def parse_datetime(value: object) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None, microsecond=0)
    text = clean_text(value)
    if not text:
        return None
    text = re.sub(r"\(UTC\+?0\)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\(UTC\)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip()
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y%m%d",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def normalize_event_time(exchange: str, dt: datetime) -> datetime:
    if exchange == "binance":
        return dt + BINANCE_TIME_OFFSET
    return dt


def serialize_decimal(value: Decimal | None) -> str | None:
    if value is None:
        return None
    return format(value.normalize(), "f") if value != value.to_integral() else format(value.quantize(Decimal("1")), "f")


def serialize_datetime(value: datetime | None) -> str | None:
    if value is None:
        return None
    return value.strftime("%Y-%m-%d %H:%M:%S")


def row_values(row: Iterable[object]) -> list[str]:
    return [clean_text(cell) for cell in row]


def build_header_map(header: list[str]) -> dict[str, int]:
    return {normalize_key(name): idx for idx, name in enumerate(header) if clean_text(name)}


def get_by_alias(values: list[object], header_map: dict[str, int], aliases: list[str]) -> object | None:
    for alias in aliases:
        idx = header_map.get(normalize_key(alias))
        if idx is not None and idx < len(values):
            return values[idx]
    return None


def find_header_row(ws, required_tokens: list[str], max_scan: int = 80) -> tuple[int, list[str]]:
    best_row = 1
    best_values: list[str] = []
    best_score = -1
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        values = row_values(row)
        joined = " ".join(v for v in values if v)
        if not joined:
            continue
        score = sum(1 for token in required_tokens if token.lower() in joined.lower())
        if score > best_score:
            best_row = row_idx
            best_values = values
            best_score = score
        if score >= len(required_tokens):
            return row_idx, values
    return best_row, best_values


def is_onchain_txid(txid: str) -> bool:
    txid = clean_text(txid).lower()
    if not txid:
        return False
    if "*" in txid:
        return False
    if txid.startswith("bcode"):
        return False
    if "gatecode" in txid:
        return False
    return len(txid) >= 20


def quantity_tolerance(asset: str, quantity: Decimal) -> Decimal:
    q = abs(quantity)
    if asset in STABLECOINS:
        pct = Decimal("0.01")
        if q >= Decimal("1000"):
            floor = Decimal("5")
        elif q >= Decimal("100"):
            floor = Decimal("1")
        elif q >= Decimal("1"):
            floor = Decimal("0.1")
        else:
            floor = Decimal("0.01")
    else:
        pct = Decimal("0.015")
        if q >= Decimal("100"):
            floor = Decimal("0.01")
        elif q >= Decimal("1"):
            floor = Decimal("0.001")
        else:
            floor = Decimal("0.0001")
    return max(floor, q * pct)


def parse_binance(path: Path) -> list[Event]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    events: list[Event] = []
    try:
        if "Deposit History 充值记录" in wb.sheetnames:
            ws = wb["Deposit History 充值记录"]
            header = row_values(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            header_map = build_header_map(header)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                status = clean_text(get_by_alias(row, header_map, ["Status 状态"]))
                if status not in {"Completed", "Success"}:
                    continue
                asset = clean_text(get_by_alias(row, header_map, ["Currency 币种"])).upper()
                quantity = parse_decimal(get_by_alias(row, header_map, ["Amount 数额"]))
                raw_time = clean_text(get_by_alias(row, header_map, ["Create Time 时间"]))
                dt = parse_datetime(raw_time)
                if not asset or quantity is None or dt is None:
                    continue
                txid = clean_text(get_by_alias(row, header_map, ["TXID 交易哈希", "txId 交易哈希"]))
                counterparty = clean_text(get_by_alias(row, header_map, ["CounterParty ID 内部划转方ID"]))
                onchain = is_onchain_txid(txid) and not counterparty
                events.append(
                    Event(
                        exchange="binance",
                        source_file=path.name,
                        source_sheet=ws.title,
                        source_row=row_idx,
                        event_class="chain_transfer",
                        direction="chain_deposit",
                        asset=asset,
                        quantity=quantity,
                        time=normalize_event_time("binance", dt),
                        raw_time=raw_time,
                        account_id=clean_text(get_by_alias(row, header_map, ["User ID 用户ID"])),
                        status=status,
                        txid=txid,
                        address=clean_text(get_by_alias(row, header_map, ["Deposit Address 接收地址"])),
                        counterparty=clean_text(get_by_alias(row, header_map, ["Source Address 发送地址"])),
                        onchain=onchain,
                        match_eligible=onchain,
                        note=None if onchain else "internal_or_masked_transfer",
                    )
                )

        if "Withdrawal History 提现记录" in wb.sheetnames:
            ws = wb["Withdrawal History 提现记录"]
            header = row_values(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            header_map = build_header_map(header)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                status = clean_text(get_by_alias(row, header_map, ["Status 状态"]))
                if status not in {"Completed", "Success"}:
                    continue
                asset = clean_text(get_by_alias(row, header_map, ["Currency 币种"])).upper()
                quantity = parse_decimal(get_by_alias(row, header_map, ["Amount 数额"]))
                raw_time = clean_text(get_by_alias(row, header_map, ["Apply Time 申请时间"]))
                dt = parse_datetime(raw_time)
                if not asset or quantity is None or dt is None:
                    continue
                txid = clean_text(get_by_alias(row, header_map, ["txId 交易哈希", "TXID 交易哈希"]))
                counterparty = clean_text(get_by_alias(row, header_map, ["CounterParty ID 内部划转方ID"]))
                onchain = is_onchain_txid(txid) and not counterparty
                events.append(
                    Event(
                        exchange="binance",
                        source_file=path.name,
                        source_sheet=ws.title,
                        source_row=row_idx,
                        event_class="chain_transfer",
                        direction="chain_withdrawal",
                        asset=asset,
                        quantity=quantity,
                        time=normalize_event_time("binance", dt),
                        raw_time=raw_time,
                        account_id=clean_text(get_by_alias(row, header_map, ["User ID 用户ID"])),
                        status=status,
                        txid=txid,
                        address=clean_text(get_by_alias(row, header_map, ["Destination Address  目标地址"])),
                        onchain=onchain,
                        match_eligible=onchain,
                        note=None if onchain else "internal_or_masked_transfer",
                    )
                )

        if "P2P 法币交易" in wb.sheetnames:
            ws = wb["P2P 法币交易"]
            header = row_values(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            header_map = build_header_map(header)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                status = clean_text(get_by_alias(row, header_map, ["Status 状态"]))
                if status != "Completed":
                    continue
                raw_direction = clean_text(get_by_alias(row, header_map, ["Buy or Sell 买卖"]))
                if raw_direction == "Taker Buy":
                    direction = "fiat_buy"
                elif raw_direction == "Taker Sell":
                    direction = "fiat_sell"
                else:
                    continue
                asset = clean_text(get_by_alias(row, header_map, ["Crypto 币种"])).upper()
                quantity = parse_decimal(get_by_alias(row, header_map, ["Amount 数额"]))
                fiat_currency = clean_text(get_by_alias(row, header_map, ["Fiat Currency 法币"])).upper()
                fiat_amount = parse_decimal(get_by_alias(row, header_map, ["Total Amount 总数额"]))
                raw_time = (
                    clean_text(get_by_alias(row, header_map, ["Release time 放币时间"]))
                    or clean_text(get_by_alias(row, header_map, ["Update time 更新时间"]))
                    or clean_text(get_by_alias(row, header_map, ["Create Time 创建时间"]))
                )
                dt = parse_datetime(raw_time)
                if not asset or quantity is None or dt is None:
                    continue
                events.append(
                    Event(
                        exchange="binance",
                        source_file=path.name,
                        source_sheet=ws.title,
                        source_row=row_idx,
                        event_class="fiat_trade",
                        direction=direction,
                        asset=asset,
                        quantity=quantity,
                        time=normalize_event_time("binance", dt),
                        raw_time=raw_time,
                        account_id=clean_text(get_by_alias(row, header_map, ["Target UID 目标用户ID"])),
                        fiat_currency=fiat_currency or None,
                        fiat_amount=fiat_amount,
                        status=status,
                        counterparty=clean_text(get_by_alias(row, header_map, ["Take ID 对手方用户ID"])),
                        raw_direction=raw_direction,
                    )
                )
    finally:
        wb.close()
    return events


def parse_gate(path: Path) -> list[Event]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    events: list[Event] = []
    try:
        deposit_sheet = "充币记录" if "充币记录" in wb.sheetnames else None
        if deposit_sheet:
            ws = wb[deposit_sheet]
            header_row, _ = find_header_row(ws, ["时间", "txid"])
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                cells = row_values(row)
                if sum(1 for cell in cells if cell) < 3:
                    continue
                dt = parse_datetime(cells[0] if len(cells) > 0 else None)
                asset = clean_text(cells[1] if len(cells) > 1 else "").upper()
                quantity = parse_decimal(cells[2] if len(cells) > 2 else None)
                txid = clean_text(cells[3] if len(cells) > 3 else "")
                if dt is None or not asset or quantity is None:
                    continue
                onchain = is_onchain_txid(txid)
                events.append(
                    Event(
                        exchange="gate",
                        source_file=path.name,
                        source_sheet=ws.title,
                        source_row=row_idx,
                        event_class="chain_transfer",
                        direction="chain_deposit",
                        asset=asset,
                        quantity=quantity,
                        time=dt,
                        raw_time=clean_text(cells[0] if len(cells) > 0 else ""),
                        txid=txid,
                        onchain=onchain,
                        match_eligible=onchain,
                        note=None if onchain else "internal_or_masked_transfer",
                    )
                )

        withdraw_sheet = "提币记录" if "提币记录" in wb.sheetnames else None
        if withdraw_sheet:
            ws = wb[withdraw_sheet]
            header_row, _ = find_header_row(ws, ["时间", "TXID", "状态"])
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                cells = row_values(row)
                if sum(1 for cell in cells if cell) < 4:
                    continue
                status = clean_text(cells[6] if len(cells) > 6 else "")
                if status and "done" not in status.lower() and "bcode" not in status.lower():
                    continue
                dt = parse_datetime(cells[1] if len(cells) > 1 else None)
                asset = clean_text(cells[2] if len(cells) > 2 else "").upper()
                quantity = parse_decimal(cells[3] if len(cells) > 3 else None)
                txid = clean_text(cells[4] if len(cells) > 4 else "")
                address = clean_text(cells[5] if len(cells) > 5 else "")
                if dt is None or not asset or quantity is None:
                    continue
                onchain = is_onchain_txid(txid) and "gatecode" not in address.lower()
                events.append(
                    Event(
                        exchange="gate",
                        source_file=path.name,
                        source_sheet=ws.title,
                        source_row=row_idx,
                        event_class="chain_transfer",
                        direction="chain_withdrawal",
                        asset=asset,
                        quantity=quantity,
                        time=dt,
                        raw_time=clean_text(cells[1] if len(cells) > 1 else ""),
                        status=status or None,
                        txid=txid,
                        address=address or None,
                        onchain=onchain,
                        match_eligible=onchain,
                        note=None if onchain else "internal_or_masked_transfer",
                    )
                )

        fiat_sheet = "法币交易" if "法币交易" in wb.sheetnames else ("Sheet3" if "Sheet3" in wb.sheetnames else None)
        if fiat_sheet:
            ws = wb[fiat_sheet]
            header_row, _ = find_header_row(ws, ["数量", "状态", "时间"])
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                cells = row_values(row)
                if sum(1 for cell in cells if cell) < 6:
                    continue
                status = clean_text(cells[5] if len(cells) > 5 else "")
                if "已完成" not in status:
                    continue
                action = clean_text(cells[7] if len(cells) > 7 else "")
                if "买入" in action:
                    direction = "fiat_buy"
                elif "卖出" in action:
                    direction = "fiat_sell"
                else:
                    continue
                pair = clean_text(cells[2] if len(cells) > 2 else "").replace(" ", "").upper()
                if "_" in pair:
                    asset, fiat_currency = pair.split("_", 1)
                else:
                    asset, fiat_currency = pair, "CNY"
                quantity = parse_decimal(cells[3] if len(cells) > 3 else None)
                fiat_amount = parse_decimal(cells[4] if len(cells) > 4 else None)
                raw_time = clean_text(cells[6] if len(cells) > 6 else "")
                dt = parse_datetime(raw_time)
                if not asset or quantity is None or dt is None:
                    continue
                events.append(
                    Event(
                        exchange="gate",
                        source_file=path.name,
                        source_sheet=ws.title,
                        source_row=row_idx,
                        event_class="fiat_trade",
                        direction=direction,
                        asset=asset,
                        quantity=quantity,
                        time=dt,
                        raw_time=raw_time,
                        fiat_currency=fiat_currency or None,
                        fiat_amount=fiat_amount,
                        status=status,
                        raw_direction=action,
                    )
                )
    finally:
        wb.close()
    return events


def parse_okx(path: Path) -> list[Event]:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    events: list[Event] = []
    try:
        if "法币交易记录" in wb.sheetnames:
            ws = wb["法币交易记录"]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                cells = row_values(row)
                if sum(1 for cell in cells if cell) < 6:
                    continue
                raw_direction = clean_text(cells[12] if len(cells) > 12 else "")
                if raw_direction == "买":
                    direction = "fiat_buy"
                elif raw_direction == "卖":
                    direction = "fiat_sell"
                else:
                    continue
                asset = clean_text(cells[2] if len(cells) > 2 else "").upper()
                quantity = parse_decimal(cells[5] if len(cells) > 5 else None)
                fiat_amount = parse_decimal(cells[6] if len(cells) > 6 else None)
                fiat_currency = clean_text(cells[3] if len(cells) > 3 else "").upper()
                raw_time = clean_text(cells[7] if len(cells) > 7 else "") or clean_text(cells[0] if len(cells) > 0 else "")
                dt = parse_datetime(raw_time)
                if not asset or quantity is None or dt is None:
                    continue
                events.append(
                    Event(
                        exchange="okx",
                        source_file=path.name,
                        source_sheet=ws.title,
                        source_row=row_idx,
                        event_class="fiat_trade",
                        direction=direction,
                        asset=asset,
                        quantity=quantity,
                        time=dt,
                        raw_time=raw_time,
                        account_id=clean_text(cells[1] if len(cells) > 1 else ""),
                        fiat_currency=fiat_currency or None,
                        fiat_amount=fiat_amount,
                        status=clean_text(cells[7] if len(cells) > 7 else "") or None,
                        raw_direction=raw_direction,
                    )
                )

        if "充币记录" in wb.sheetnames:
            ws = wb["充币记录"]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                cells = row_values(row)
                if sum(1 for cell in cells if cell) < 6:
                    continue
                dt = parse_datetime(cells[5] if len(cells) > 5 else None)
                asset = clean_text(cells[1] if len(cells) > 1 else "").upper()
                quantity = parse_decimal(cells[3] if len(cells) > 3 else None)
                txid = clean_text(cells[4] if len(cells) > 4 else "")
                if dt is None or not asset or quantity is None:
                    continue
                onchain = is_onchain_txid(txid)
                events.append(
                    Event(
                        exchange="okx",
                        source_file=path.name,
                        source_sheet=ws.title,
                        source_row=row_idx,
                        event_class="chain_transfer",
                        direction="chain_deposit",
                        asset=asset,
                        quantity=quantity,
                        time=dt,
                        raw_time=clean_text(cells[5] if len(cells) > 5 else ""),
                        account_id=clean_text(cells[0] if len(cells) > 0 else ""),
                        txid=txid,
                        address=clean_text(cells[2] if len(cells) > 2 else "") or None,
                        onchain=onchain,
                        match_eligible=onchain,
                        note=None if onchain else "internal_or_masked_transfer",
                    )
                )

        if "提币记录" in wb.sheetnames:
            ws = wb["提币记录"]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                cells = row_values(row)
                if sum(1 for cell in cells if cell) < 6:
                    continue
                dt = parse_datetime(cells[5] if len(cells) > 5 else None)
                asset = clean_text(cells[1] if len(cells) > 1 else "").upper()
                quantity = parse_decimal(cells[3] if len(cells) > 3 else None)
                txid = clean_text(cells[4] if len(cells) > 4 else "")
                if dt is None or not asset or quantity is None:
                    continue
                onchain = is_onchain_txid(txid)
                events.append(
                    Event(
                        exchange="okx",
                        source_file=path.name,
                        source_sheet=ws.title,
                        source_row=row_idx,
                        event_class="chain_transfer",
                        direction="chain_withdrawal",
                        asset=asset,
                        quantity=quantity,
                        time=dt,
                        raw_time=clean_text(cells[5] if len(cells) > 5 else ""),
                        account_id=clean_text(cells[0] if len(cells) > 0 else ""),
                        txid=txid,
                        address=clean_text(cells[2] if len(cells) > 2 else "") or None,
                        onchain=onchain,
                        match_eligible=onchain,
                        note=None if onchain else "internal_or_masked_transfer",
                    )
                )
    finally:
        wb.close()
    return events


def parse_all_events(input_dir: Path = DEFAULT_INPUT_DIR) -> tuple[list[Event], dict[str, object]]:
    events: list[Event] = []
    warnings_list: list[str] = []
    workbook_counts: dict[str, int] = {}

    for exchange_dir in sorted(p for p in input_dir.iterdir() if p.is_dir()):
        parser = {
            "binance": parse_binance,
            "gate": parse_gate,
            "okx": parse_okx,
        }.get(exchange_dir.name)
        if parser is None:
            continue
        for workbook in sorted(exchange_dir.glob("*.xlsx")):
            parsed = parser(workbook)
            workbook_counts[str(workbook.relative_to(ROOT_DIR))] = len(parsed)
            events.extend(parsed)

    events.sort(key=lambda item: (item.time, item.exchange, item.source_file, item.source_sheet, item.source_row))
    for idx, event in enumerate(events, start=1):
        event.event_id = f"evt_{idx:06d}"

    metadata = {
        "warnings": warnings_list,
        "workbook_event_counts": workbook_counts,
        "ignored_sources": [
            "binance/*/OTC Trading records are not included in the main matching flow because they are closer to channel exchange records and can introduce false matches.",
            "okx account bills are not used as separate fiat event sources; only fiat trade, deposit, and withdrawal records are kept as main event sources.",
        ],
    }
    return events, metadata


def prepare_event_indices(events: list[Event]) -> tuple[dict[tuple[str, str], list[Event]], dict[tuple[str, str], list[datetime]]]:
    by_key: dict[tuple[str, str], list[Event]] = defaultdict(list)
    for event in events:
        if event.match_eligible:
            by_key[(event.direction, event.asset)].append(event)
    time_index: dict[tuple[str, str], list[datetime]] = {}
    for key, items in by_key.items():
        items.sort(key=lambda item: item.time)
        time_index[key] = [item.time for item in items]
    return by_key, time_index


def get_candidate_pool(
    anchor: Event,
    counter_direction: str,
    mode: str,
    by_key: dict[tuple[str, str], list[Event]],
    time_index: dict[tuple[str, str], list[datetime]],
) -> list[Event]:
    key = (counter_direction, anchor.asset)
    candidates = by_key.get(key, [])
    if not candidates:
        return []
    times = time_index[key]
    if mode == "forward":
        left = bisect_left(times, anchor.time)
        right = bisect_right(times, anchor.time + MATCH_WINDOW)
    else:
        left = bisect_left(times, anchor.time - MATCH_WINDOW)
        right = bisect_right(times, anchor.time)

    tolerance = quantity_tolerance(anchor.asset, anchor.quantity)
    max_quantity = anchor.quantity + tolerance * Decimal("2")
    sliced = [
        item
        for item in candidates[left:right]
        if item.event_id != anchor.event_id and item.quantity <= max_quantity
    ]
    sliced.sort(key=lambda item: abs((item.time - anchor.time).total_seconds()))
    return sliced[:MAX_NEARBY_CANDIDATES]


def score_candidate_group(anchor: Event, group: list[Event]) -> tuple[Decimal, float]:
    total_quantity = sum(item.quantity for item in group)
    gap = abs(total_quantity - anchor.quantity)
    farthest_gap_days = max(abs((item.time - anchor.time).total_seconds()) for item in group) / 86400
    return gap, farthest_gap_days


def confidence_level(anchor: Event, gap: Decimal, farthest_gap_days: float) -> str:
    tolerance = quantity_tolerance(anchor.asset, anchor.quantity)
    if gap <= tolerance and farthest_gap_days <= 7:
        return "high"
    if gap <= tolerance * Decimal("1.5") and farthest_gap_days <= 30:
        return "medium"
    return "low"


def build_anchor_matches(
    anchor: Event,
    counter_direction: str,
    mode: str,
    label_type: str,
    candidate_pool: list[Event],
) -> list[dict[str, object]]:
    if not candidate_pool:
        return []

    chronological = sorted(candidate_pool, key=lambda item: item.time)
    tolerance = quantity_tolerance(anchor.asset, anchor.quantity)
    max_gap = tolerance * Decimal("2")
    best_groups: dict[tuple[str, ...], dict[str, object]] = {}

    def consider(group: list[Event]) -> None:
        gap, farthest_gap_days = score_candidate_group(anchor, group)
        if gap > max_gap:
            return
        key = tuple(item.event_id for item in group)
        total_quantity = sum(item.quantity for item in group)
        fiat_ids = [anchor.event_id] if anchor.event_class == "fiat_trade" else [item.event_id for item in group if item.event_class == "fiat_trade"]
        chain_ids = [anchor.event_id] if anchor.event_class == "chain_transfer" else [item.event_id for item in group if item.event_class == "chain_transfer"]
        if anchor.event_class == "fiat_trade":
            chain_ids = [item.event_id for item in group]
        if anchor.event_class == "chain_transfer":
            fiat_ids = [item.event_id for item in group]
        candidate_record = {
            "label_type": label_type,
            "anchor_event_id": anchor.event_id,
            "anchor_direction": anchor.direction,
            "candidate_event_ids": [item.event_id for item in group],
            "candidate_count": len(group),
            "relation_shape": "one_to_one" if len(group) == 1 else "one_to_many",
            "asset": anchor.asset,
            "anchor_quantity": serialize_decimal(anchor.quantity),
            "candidate_total_quantity": serialize_decimal(total_quantity),
            "quantity_gap": serialize_decimal(gap),
            "confidence": confidence_level(anchor, gap, farthest_gap_days),
            "anchor_time": serialize_datetime(anchor.time),
            "candidate_time_start": serialize_datetime(min(item.time for item in group)),
            "candidate_time_end": serialize_datetime(max(item.time for item in group)),
            "farthest_gap_days": round(farthest_gap_days, 4),
            "fiat_event_ids": fiat_ids,
            "chain_event_ids": chain_ids,
        }
        score = (
            float(gap / max(abs(anchor.quantity), Decimal("1")))
            + farthest_gap_days / 120
            + max(0, len(group) - 1) * 0.03
        )
        candidate_record["_score"] = score
        previous = best_groups.get(key)
        if previous is None or score < previous["_score"]:
            best_groups[key] = candidate_record

    for item in candidate_pool:
        consider([item])

    for start in range(len(chronological)):
        running: list[Event] = []
        running_total = Decimal("0")
        for idx in range(start, min(len(chronological), start + MAX_SEQUENCE_LENGTH)):
            running.append(chronological[idx])
            running_total += chronological[idx].quantity
            if len(running) >= 2:
                consider(running.copy())
            if running_total > anchor.quantity + max_gap:
                break

    ranked = sorted(best_groups.values(), key=lambda item: item["_score"])
    output = []
    for item in ranked[:MAX_GROUPS_PER_ANCHOR]:
        item.pop("_score", None)
        output.append(item)
    return output


def build_labels(events: list[Event]) -> dict[str, list[dict[str, object]]]:
    by_key, time_index = prepare_event_indices(events)
    labels: dict[str, list[dict[str, object]]] = {
        "fiat_buy_to_chain_withdrawal": [],
        "chain_withdrawal_to_fiat_buy": [],
        "fiat_sell_to_chain_deposit": [],
        "chain_deposit_to_fiat_sell": [],
    }

    for anchor in events:
        if not anchor.match_eligible:
            continue
        config = MATCH_CONFIG.get(anchor.direction)
        if config is None:
            continue
        pool = get_candidate_pool(anchor, config["counter_direction"], config["mode"], by_key, time_index)
        matches = build_anchor_matches(anchor, config["counter_direction"], config["mode"], config["label_type"], pool)
        labels[config["label_type"]].extend(matches)
    return labels


def summarize(events: list[Event], labels: dict[str, list[dict[str, object]]]) -> dict[str, object]:
    direction_counts = Counter(event.direction for event in events)
    eligible_counts = Counter(event.direction for event in events if event.match_eligible)
    onchain_excluded = Counter(
        event.direction for event in events if event.event_class == "chain_transfer" and not event.match_eligible
    )
    label_counts = {label_type: len(items) for label_type, items in labels.items()}
    return {
        "total_events": len(events),
        "direction_counts": dict(direction_counts),
        "eligible_direction_counts": dict(eligible_counts),
        "excluded_non_onchain_counts": dict(onchain_excluded),
        "label_counts": label_counts,
    }


def event_to_json(event: Event) -> dict[str, object]:
    return {
        "event_id": event.event_id,
        "exchange": event.exchange,
        "source_file": event.source_file,
        "source_sheet": event.source_sheet,
        "source_row": event.source_row,
        "event_class": event.event_class,
        "direction": event.direction,
        "asset": event.asset,
        "quantity": serialize_decimal(event.quantity),
        "time": serialize_datetime(event.time),
        "raw_time": event.raw_time,
        "account_id": event.account_id,
        "fiat_currency": event.fiat_currency,
        "fiat_amount": serialize_decimal(event.fiat_amount),
        "status": event.status,
        "txid": event.txid,
        "address": event.address,
        "counterparty": event.counterparty,
        "onchain": event.onchain,
        "match_eligible": event.match_eligible,
        "raw_direction": event.raw_direction,
        "note": event.note,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build normalized fiat/on-chain label tags from exchange evidence workbooks.")
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR, help="Input root containing exchange workbook folders.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output JSON path.")
    args = parser.parse_args()

    events, metadata = parse_all_events(args.input_dir)
    labels = build_labels(events)
    payload = {
        "generated_at": serialize_datetime(datetime.now()),
        "source_root": str(args.input_dir.relative_to(ROOT_DIR)),
        "assumptions": {
            "binance_time_shift": "Binance sheet timestamps are normalized by +8 hours to align with Gate/OKX local time and cross-exchange txid samples.",
            "matching_window_days": MATCH_WINDOW.days,
            "amount_tolerance": "Tolerance depends on asset and quantity; stablecoins default to about 1% with an absolute floor.",
            "matching_scope": "Only completed P2P/法币成交 records and on-chain-looking charge/withdraw records enter matching.",
        },
        "metadata": metadata,
        "summary": summarize(events, labels),
        "events": [event_to_json(event) for event in events],
        "labels": labels,
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)

    print(f"Wrote {args.output}")
    print(json.dumps(payload["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
